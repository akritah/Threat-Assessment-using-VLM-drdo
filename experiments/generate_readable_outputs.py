import os
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = r"c:\Drdo threat detection"
CSV_PATH = os.path.join(PROJECT_ROOT, "evaluation", "csv", "evaluation_results.csv")
REPORTS_DIR = os.path.join(PROJECT_ROOT, "evaluation", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(REPORTS_DIR, "evaluation_results.xlsx")
HTML_PATH = os.path.join(REPORTS_DIR, "evaluation_results.html")

def build_excel():
    print("Generating styled Excel sheet...")
    if not os.path.exists(CSV_PATH):
        print(f"Error: CSV file not found at {CSV_PATH}")
        return
        
    # Load data
    df = pd.read_csv(CSV_PATH)
    
    # Save as basic excel first
    df.to_excel(EXCEL_PATH, index=False)
    
    # Load using openpyxl for styling
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.title = "VLM Evaluations"
    
    # Define styles
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Threat level fills
    fill_high = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")  # Soft red
    fill_med = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Soft yellow
    fill_low = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Soft green
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Style header row
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    # Style data rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 65  # Give cells breathing room
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            # Default alignments
            if col_idx in [1, 2, 3, 9, 10]:  # Video ID, Category, Frames, Threat Level, Latency
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Color-code threat levels (column 9)
            if col_idx == 9:
                val = str(cell.value).strip().lower()
                if "high" in val:
                    cell.fill = fill_high
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
                elif "medium" in val:
                    cell.fill = fill_med
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="7F6000")
                elif "low" in val:
                    cell.fill = fill_low
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="375623")
                    
    # Adjust column widths
    column_widths = {
        1: 22,  # Video ID
        2: 18,  # Ground Truth Category
        3: 12,  # Frames Used
        4: 45,  # Base Gemma Output
        5: 35,  # Fine-Tuned Gemma Output
        6: 12,  # Video-LLaVA Output
        7: 25,  # Predicted Activity
        8: 55,  # Threat Assessment
        9: 15,  # Threat Level
        10: 22, # Inference Time
        11: 20  # Notes
    }
    
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
    wb.save(EXCEL_PATH)
    print(f"Excel saved successfully to: {EXCEL_PATH}")

def build_html():
    print("Generating styled HTML webpage...")
    if not os.path.exists(CSV_PATH):
        return
        
    records = []
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for r in reader:
            records.append(r)
            
    html_content = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>VLM Threat Assessment Evaluations</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f7f9fa;
            color: #333;
            margin: 0;
            padding: 20px;
        }
        h1 {
            color: #2c3e50;
            text-align: center;
            margin-bottom: 20px;
        }
        .container {
            max-width: 98%;
            margin: 0 auto;
            background: #fff;
            padding: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            border-radius: 8px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #e0e0e0;
        }
        th {
            background-color: #34495e;
            color: #fff;
            font-weight: 600;
            position: sticky;
            top: 0;
        }
        tr:hover {
            background-color: #f5f6fa;
        }
        .badge {
            display: inline-block;
            padding: 6px 12px;
            font-weight: bold;
            border-radius: 4px;
            text-align: center;
            font-size: 0.9em;
        }
        .badge-high { background-color: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
        .badge-med { background-color: #fff3cd; color: #856404; border: 1px solid #ffeeba; }
        .badge-low { background-color: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        
        .text-cell {
            max-width: 320px;
            max-height: 120px;
            overflow-y: auto;
            white-space: pre-wrap;
            font-size: 0.88em;
            background-color: #fafbfc;
            border: 1px solid #f1f2f6;
            padding: 8px;
            border-radius: 4px;
        }
        .vid-cell { font-weight: bold; color: #2980b9; }
        .latency-cell { font-family: monospace; font-size: 0.85em; color: #555; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 VLM Surveillance Threat Assessment Logs (100 Videos)</h1>
        <table>
            <thead>
                <tr>
                    <th>Video ID</th>
                    <th>Category</th>
                    <th>FT Activity Caption</th>
                    <th>VLM Reasoning & Assessment</th>
                    <th>Threat Level</th>
                    <th>Inference Time</th>
                </tr>
            </thead>
            <tbody>
    """
    
    for r in records:
        vid = r.get("Video ID", "")
        cat = r.get("Ground Truth Category", "")
        ft_caption = r.get("Fine-Tuned Gemma Output", "")
        reasoning = r.get("Threat Assessment", "")
        level = r.get("Threat Level", "Low").strip()
        inf_time = r.get("Inference Time", "")
        
        # Determine badge style
        l_lower = level.lower()
        if "high" in l_lower:
            badge_class = "badge badge-high"
        elif "medium" in l_lower:
            badge_class = "badge badge-med"
        else:
            badge_class = "badge badge-low"
            
        html_content += f"""
                <tr>
                    <td class="vid-cell">{vid}</td>
                    <td>{cat}</td>
                    <td><div class="text-cell">{ft_caption}</div></td>
                    <td><div class="text-cell">{reasoning}</div></td>
                    <td><span class="{badge_class}">{level}</span></td>
                    <td class="latency-cell">{inf_time}</td>
                </tr>
        """
        
    html_content += """
            </tbody>
        </table>
    </div>
</body>
</html>
    """
    
    with open(HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML Report saved successfully to: {HTML_PATH}")

def main():
    build_excel()
    build_html()

if __name__ == "__main__":
    main()
